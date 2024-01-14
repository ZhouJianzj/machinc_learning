"""
将之前的爬取到的天气数据存入到excel表格中
"""
import weather
import openpyxl

excel = openpyxl.Workbook()
sheet1 = excel.create_sheet("各景点天气")

# 发情请求
data = weather.get_request()
# 数据转换
parse_data = weather.data_parse(data)
print(parse_data)
# 数据入sheet
for item in parse_data:
    sheet1.append(item)
# 保存工作簿
excel.save("数据表.xlsx")

print("*" * 150)
# 加载excel表格并操作数据
workboot = openpyxl.load_workbook("数据表.xlsx")
# 获取工作簿中所有表格名称
print(workboot.sheetnames)
# 获取指定的表格
sheet1 = workboot["各景点天气"]
# 遍历每一行数据
lst = []
# 获取每一行
for row in sheet1.rows:
    row_lst = []
    # 获取一行中每一个单元格的数据
    for cell in row:
        row_lst.append(cell.value)
    lst.append(row_lst)
print(lst)
for item in lst:
    print(item)